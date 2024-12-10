# import time
# import pandas as pd
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from webdriver_manager.chrome import ChromeDriverManager
# from selenium.webdriver.chrome.service import Service
# from selenium import webdriver
# import os

# class CurrencyTest:
#     def __init__(self, driver, url):
#         self.service = Service(ChromeDriverManager().install())
#         self.driver = webdriver.Chrome(service=self.service)
#         self.url = url
#         self.results = []

#     def test_currency_filter(self):
#         """
#         Perform currency filtering and ensure property tiles currency changes according to the selected currency.
#         """
#         try:
#             self.driver.get(self.url)
#             print("Page loaded")

#             # Locate the currency dropdown element
#             try:
#                 currency_dropdown = WebDriverWait(self.driver, 20).until(
#                     EC.presence_of_element_located((By.ID, 'js-currency-sort-footer'))
#                 )
#                 print("Currency dropdown found:", currency_dropdown)
#             except Exception as e:
#                 print(f"Error locating currency dropdown: {e}")
#                 raise

#             # Ensure the dropdown is active before selecting an option
#             try:
#                 WebDriverWait(self.driver, 20).until(
#                     EC.element_to_be_clickable(currency_dropdown)
#                 )
#                 print("Currency dropdown is active")
#             except Exception as e:
#                 print(f"Error waiting for dropdown activation: {e}")
#                 raise

#             # Locate the currency options inside the `.select-ul` <ul> element
#             try:
#                 currency_options = currency_dropdown.find_elements(By.XPATH, './/ul[@class="select-ul"]/li')
#                 print(f"Found {len(currency_options)} currency options.")
#             except Exception as e:
#                 print(f"Error locating currency options: {e}")
#                 raise

#             # Iterate through each currency option
#             for option in currency_options:
#                 try:
#                     # Ensure the option is visible before proceeding
#                     print("Before visibility check...")
#                     WebDriverWait(self.driver, 10).until(
#                         EC.visibility_of(option)
#                     )
#                     print("After visibility check...")

#                     # Try to get the currency text from the <p> tag inside each <li> option
#                     try:
#                         # Correct XPath to get the text inside the <p> tag
#                         currency_text = option.find_element(By.XPATH, './/div[@class="option"]/p').text.strip()
#                         if not currency_text:
#                             raise Exception("Currency text is empty")
#                         print(f"Currency found: {currency_text}")
#                     except Exception as e:
#                         print(f"Error getting currency text from option: {e}")
#                         continue  # Skip this option if we can't get the text

#                     # Scroll the option into view
#                     self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", option)

#                     # Wait for the option to be clickable
#                     WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(option))
#                     print(f"Clicking on currency: {currency_text}")
                    
#                     # Click the currency option
#                     option.click()

#                     # Wait for the page to reload or update
#                     time.sleep(5)

#                     # Check property tiles for currency changes
#                     property_prices = self.driver.find_elements(By.CLASS_NAME, 'js-price-value')  # Adjusted class name
#                     mismatched_prices = [
#                         price.get_attribute('title') 
#                         for price in property_prices 
#                         if currency_text not in price.text
#                     ]

#                     # Collect results
#                     self.results.append({
#                         'currency': currency_text,
#                         'status': 'Pass' if not mismatched_prices else 'Fail',
#                         'mismatched': mismatched_prices
#                     })

#                 except Exception as e:
#                     print(f"Error during currency option interaction: {e}")
#                     self.results.append({
#                         'currency': 'N/A',
#                         'status': 'Fail',
#                         'comments': f"Error interacting with currency option: {str(e)}"
#                     })

#         except Exception as e:
#             print(f"Error during currency filtering: {e}")
#             self.results.append({
#                 'currency': 'N/A',
#                 'status': 'Fail',
#                 'comments': f'Error during currency filtering: {str(e)}'
#             })

#     def generate_excel_report(self, filename="reports/currency_test_report.xlsx"):
#         """
#         Generate an Excel report from the test results.
#         """
#         # Ensure the 'reports' directory exists
#         os.makedirs(os.path.dirname(filename), exist_ok=True)

#         data = []
#         for result in self.results:
#             data.append({
#                 "Currency": result.get("currency"),
#                 "Status": result.get("status"),
#                 "Comments": result.get("comments", "No comments"),
#                 "Mismatched Prices": ', '.join(result.get("mismatched", [])) if "mismatched" in result else "N/A"
#             })

#         # Convert to a DataFrame
#         df = pd.DataFrame(data)

#         # Save to an Excel file
#         df.to_excel(filename, index=False)
#         print(f"Report saved as {filename}")


# if __name__ == "__main__":
#     # Setup WebDriver using Service
#     service = Service(ChromeDriverManager().install())
#     driver = webdriver.Chrome(service=service)
#     url = "https://www.alojamiento.io/"  # Replace with the actual website URL

#     try:
#         # Run the test
#         currency_test = CurrencyTest(driver, url)
#         currency_test.test_currency_filter()

#         # Generate the report
#         currency_test.generate_excel_report()

#     finally:
#         driver.quit()


# Correct code
import time
import os
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class CurrencySelectionBot:
    def __init__(self, url, log_callback=None):
        self.chrome_options = Options()
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_argument("--disable-extensions")
        self.chrome_options.add_argument("--disable-gpu")
        self.chrome_options.add_argument("--no-sandbox")
        self.service = Service(ChromeDriverManager().install())
        self.driver = None
        self.url = url
        self.log_callback = log_callback
        self.results = []

    def log(self, message):
        print(message)
        if self.log_callback:
            self.log_callback(message)

    def setup_driver(self):
        self.driver = webdriver.Chrome(service=self.service, options=self.chrome_options)
        self.wait = WebDriverWait(self.driver, 30)

    def run_currency_selection_test(self):
        try:
            self.setup_driver()
            self.log("🌐 Navigating to website...")
            self.driver.get(self.url)

            self.log("🔍 Searching for currency dropdown...")
            currency_dropdown = self.wait.until(
                EC.presence_of_element_located((By.ID, 'js-currency-sort-footer'))
            )
            self.log("✅ Currency dropdown found!")

            currency_options = currency_dropdown.find_elements(
                By.XPATH, './/ul[@class="select-ul"]/li'
            )
            self.log(f"💰 Found {len(currency_options)} currency options")

            for index, option in tqdm(enumerate(currency_options, 1),
                                      total=len(currency_options),
                                      desc="Processing currencies",
                                      ncols=100, unit="option"):
                try:
                    # Extract the currency name
                    currency_raw_text = option.get_attribute('innerText')
                    # print(currency_raw_text)
                    currency_match = re.search(r'\((.*?)\)', currency_raw_text)
                    currency_text = currency_match.group(1)
                    self.log(f"\n🔄 Processing Currency Option {index}: {currency_text}")
                    
                    # Capture initial prices
                    property_prices = self.driver.find_elements(By.CLASS_NAME, 'js-price-value')
                    initial_prices = [
                        re.sub(r'^.{3}', '', price.text.strip()) for price in property_prices
                    ]
                    # initial_text = currency_text
                    self.driver.execute_script("arguments[0].click();", option)
                    time.sleep(3)  # Wait for prices to update

                    # Capture updated prices
                    updated_prices = [
                        re.sub(r'^.{3}', '', price.text.strip()) for price in property_prices
                    ]

                    # Compare initial and updated prices
                    price_changes = []
                    status = 'Pass'
                    for initial, updated in zip(initial_prices, updated_prices):
                        if initial != updated:
                            print(f"Initial : {initial:<10} --- Updated : {updated:<10}")
                            # price_changes.append(f"{initial} ➡ {updated}")
                        else:
                            status = 'Fail'  # Mark as fail if any price does not change

                    if status == 'Pass':
                        comments = f"Prices updated successfully in {currency_text}"
                        self.log(f"🟢 Property prices successfully updated in {currency_text} ")
                    else:
                        comments = "Prices did not update"
                        self.log(f"❌ Prices did not update for {currency_text}")
                    
                    self.results.append({
                        'url': self.url,
                        'currency': currency_text,
                        'status': status,
                        'comments': comments
                    })

                except Exception as e:
                    self.log(f"❌ Error with currency option {index}: {e}")
                    self.results.append({
                        'url': self.url,
                        'currency': "Unknown",
                        'status': 'Fail',
                        'comments': f"Error: {e}"
                    })
        except Exception as e:
            self.log(f"❌ Critical Test Error: {e}")
            return False
        finally:
            if self.driver:
                self.driver.quit()
        return True

    def generate_excel_report(self):
        try:
            os.makedirs('reports', exist_ok=True)
            report_file = 'reports/all_the_reports.xlsx'

            if os.path.exists(report_file):
                workbook = openpyxl.load_workbook(report_file)
                if "Currency" not in workbook.sheetnames:
                    sheet = workbook.create_sheet(title="Currency")
                else:
                    sheet = workbook["Currency"]
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Currency"

            headers = ['Page URL', 'Currency', 'Status', 'Comments']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row, result in enumerate(self.results, start=2):
                sheet.cell(row=row, column=1, value=result.get('url', ''))
                sheet.cell(row=row, column=2, value=result.get('currency', ''))
                sheet.cell(row=row, column=3, value=result.get('status', ''))
                sheet.cell(row=row, column=4, value=result.get('comments', ''))

            for col in range(1, 5):
                column_letter = get_column_letter(col)
                max_length = max(len(str(cell.value or '')) for cell in sheet[get_column_letter(col)])
                sheet.column_dimensions[column_letter].width = max_length + 2

            workbook.save(report_file)
            self.log(f"✅ Excel Report generated: {report_file}")
            return report_file
        except Exception as e:
            self.log(f"❌ Error generating Excel report: {e}")
            return None


def main():
    url = "https://www.alojamiento.io/property/cabrils/BC-1178728"
    bot = CurrencySelectionBot(url)
    if bot.run_currency_selection_test():
        print("✅ Currency Selection Test Completed Successfully!")
        bot.generate_excel_report()
    else:
        print("❌ Currency Selection Test Failed")

# if __name__ == "__main__":
#     main()
